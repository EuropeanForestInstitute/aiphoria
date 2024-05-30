# Load a local copy of the current ODYM branch:
import sys
import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import pickle
import openpyxl
import pylab

# Specify path to dynamic stock model and to datafile, relative
MainPath = os.path.join('..', 'odym', 'modules')
DataPath = os.path.join('..', 'docs', 'files')
sys.path.insert(0, MainPath)

# add ODYM module directory to system path, absolute
sys.path.insert(0, os.path.join(os.getcwd(), '..', 'odym', 'modules'))
sys.path.insert(0, os.path.join(os.getcwd(), '..', 'docs', 'files'))

# Import ODYM files
import ODYM_Classes as msc  # import the ODYM class file
import ODYM_Functions as msf  # import the ODYM function file
import dynamic_stock_model as dsm  # import the dynamic stock model library

# Read available years from the file
lifetimes_filename = 'testi_product_data.xlsx'
lifetimes_datasheet_name = 'Average_Lifetime'

regions = ['Argentina', 'Brazil', 'Canada',
           'Denmark', 'Ethiopia', 'France',
           'Greece', 'Hungary', 'Indonesia']

# Minimum and maximum year, inclusive
use_automatic_year_detection = True
min_year = 1960
max_year = 2009

years = []
for year in range(min_year, max_year + 1):
    years.append(year)

# Create dictionary of model classifications
ModelClassification = {}

# Classification for time labelled 'Time' must always be present, with Items containing a list of odered integers
# representing years, months, or other discrete time intervals
ModelClassification['Time'] = msc.Classification(Name='Time', Dimension='Time', ID=1, Items=years)

# Classification for cohort is used to track age-cohorts in the stock.
ModelClassification['Cohort'] = msc.Classification(Name='Age-cohort', Dimension='Time', ID=2, Items=years)

# Classification for elements labelled 'Element' must always be present, with Items containing a list of the
# symbols of the elements covered.
ModelClassification['Element'] = msc.Classification(Name='Elements', Dimension='Element', ID=3, Items=['Fe'])

# Classification for regions is chosen to include the regions that are in the scope of this analysis.
ModelClassification['Region'] = msc.Classification(Name='Regions', Dimension='Region', ID=4, Items=regions)

# Get model time start, end, and duration:
Model_Time_Start = int(min(ModelClassification['Time'].Items))
Model_Time_End = int(max(ModelClassification['Time'].Items))
Model_Duration = Model_Time_End - Model_Time_Start

IndexTable = pd.DataFrame(
    {'Aspect': ['Time', 'Age-cohort', 'Element', 'Region'],  # 'Time' and 'Element' must be present!
     'Description': ['Model aspect "time"', 'Model aspect "age-cohort"', 'Model aspect "Element"',
                     'Model aspect "Region where flow occurs"'],
     'Dimension': ['Time', 'Time', 'Element', 'Region'],  # 'Time' and 'Element' are also dimensions
     'Classification': [ModelClassification[Aspect] for Aspect in ['Time', 'Cohort', 'Element', 'Region']],
     # Unique one letter (upper or lower case) indices to be used later for calculations.
     'IndexLetter': ['t', 'c', 'e', 'r']})

# Default indexing of IndexTable, other indices are produced on the fly
IndexTable.set_index('Aspect', inplace=True)
print(IndexTable)

# Initialize MFA system
Dyn_MFA_System = msc.MFAsystem(Name='StockAccumulationSystem',
                               Geogr_Scope='9SelectedRegions',
                               Unit='kt',
                               ProcessList=[],
                               FlowDict={},
                               StockDict={},
                               ParameterDict={},
                               Time_Start=Model_Time_Start,
                               Time_End=Model_Time_End,
                               IndexTable=IndexTable,
                               Elements=IndexTable.loc['Element'].Classification.Items)

# LIFETIMES
lifetimes_workbook = openpyxl.load_workbook(os.path.join(DataPath, lifetimes_filename), data_only=True)
lifetimes_datasheet = lifetimes_workbook[lifetimes_datasheet_name]

# Use 1-based column numbers (= as column number is shown in Excel file)
lifetimes_col_region_name = 1
lifetimes_col_lifetime_value = 2

lifetimes_for_regions = []
lifetimes_rows = lifetimes_datasheet.iter_rows()
lifetime_line_number = 0
for row in lifetimes_rows:
    lifetime_line_number = lifetime_line_number + 1
    if lifetime_line_number == 1:
        # Ignore first line in file
        continue

    region_name = row[lifetimes_col_region_name - 1].value
    lifetime_value = row[lifetimes_col_lifetime_value - 1].value

    # Ignore lines that do not contain proper data
    if (region_name is None) or (lifetime_value is None):
        continue

    lifetimes_for_regions.append([region_name, lifetime_value])

for entry in lifetimes_for_regions:
    print(entry)

# Get lifetimes for the regions
# num_regions = 10
# for row in range(1, num_regions):
#     lifetimes_for_regions.append(lifetimes_datasheet.cell(row + 1, lifetimes_col_lifetime_value).value)
#
# print(lifetimes_for_regions)
